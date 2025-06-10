from typing import Any
from django.core.management.base import BaseCommand
from pprint import pprint

import openpyxl
from slugify import slugify

from main import models

HEADERS = [
    "№",
    "Баркод",
    "Название",
    "Издательство",
    "Описание",
    "Возраст",
    "Габариты см",
    "Кол-во страниц",
    "Переплёт",
    "Категория",
    "Под категория",
    "Цена",
    "Ссылка на фото",
]


def get_data_from_excel_file(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    my_list = list()
    for idx, value in enumerate(
        ws.iter_rows(
            min_row=1, max_row=ws.max_row + 1, min_col=1, max_col=13, values_only=True
        )
    ):
        is_all_none = all(list(map(lambda x: x is None, value[1:])))
        if not is_all_none:
            if idx == 0:
                continue
            my_list.append(dict(zip(HEADERS, value)))

    return my_list


class Command(BaseCommand):
    def handle(self, *args: Any, **options: Any) -> str | None:
        excel_data = get_data_from_excel_file("Список товаров.xlsx")

        for row in excel_data:
            barcode = row.get("Баркод")

            if not barcode:
                print("no barcode, continue")
                continue

            age = row.get("Возраст")
            size = row.get("Габариты см")
            publisher = row.get("Издательство")
            category = row.get("Категория")
            pages = row.get("Кол-во страниц")
            title = row.get("Название")
            description = row.get("Описание")
            binding = row.get("Переплёт")
            subcategory = row.get("Под категория")

            price = row.get("Цена")

            if category is not None:
                category, category_created = models.Category.objects.get_or_create(
                    name=category
                )
            else:
                category = None

            if subcategory is not None and category is not None:
                product_subcategory, product_subcategory_created = (
                    models.Subcategory.objects.get_or_create(
                        name=subcategory,
                        slug=slugify(subcategory),
                        category=category,
                    )
                )
            else:
                product_subcategory = None

            publisher_obj, publisher_created = (
                models.PublishingHouse.objects.get_or_create(
                    name=publisher,
                    slug=slugify(publisher),
                )
            )
            age, created_age = models.CategoryAge.objects.get_or_create(age=age)

            try:
                product = models.Product.objects.get(barcode=int(barcode))
                product.ages.add(age)
                product.size = size
                product.publisher = publisher_obj
                product.main_category = category
                product.title = title
                product.slug = slugify(title)
                product.description = description
                product.binding = binding
                product.subcategory = product_subcategory
                product.price = int(price) if price else 0
                product.pages = pages
                product.save()
                print(f"[UPDATED]: {product}")

            except models.Product.DoesNotExist:
                product = models.Product.objects.create(
                    barcode=int(barcode),
                    title=title,
                    size=size,
                    slug=slugify(title),
                    price=price,
                    description=description,
                    binding=binding,
                    main_category=category,
                    subcategory=product_subcategory,
                    publisher=publisher_obj,
                    pages=pages,
                )
                product.ages.add(age)
                print(f"[CREATED]: {product}")

            # return product, is_updated, is_created
